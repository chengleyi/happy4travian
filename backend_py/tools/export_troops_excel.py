"""兵种参数导出为 Excel（openpyxl版）

读取静态 JSON 数据，将各部落兵种参数导出到 xlsx 文件；中文映射可按内置表补充。
"""
import os
import json
from datetime import timedelta
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

CN_TRIBES = {
    1: "罗马", 2: "条顿", 3: "高卢", 4: "自然", 5: "纳塔",
    6: "埃及", 7: "匈奴", 8: "斯巴达", 9: "维京"
}
CN_UNITS = {
    1: ["军团兵","禁卫兵","帝国兵","侦察骑兵","帝国骑兵","凯撒骑兵","攻城槌","投石车","参议员","移民"],
    2: ["棒兵","矛兵","斧兵","侦察兵","圣骑士","条顿骑士","攻城槌","投石车","酋长","移民"],
    3: ["方阵","剑士","探路者","雷霆骑兵","德鲁伊骑士","海都安骑士","攻城槌","投石车","首领","移民"]
}
CN_TYPE = {"i": "步兵", "c": "骑兵"}

def fmt_time(sec):
    """秒转时分秒字符串"""
    if not isinstance(sec, (int, float)) or sec <= 0:
        return "0:00:00"
    return str(timedelta(seconds=int(sec)))

def export_excel(json_path: str, out_path: str):
    """从 JSON 生成 xlsx 并返回输出路径"""
    with open(json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["版本 version", data.get("version"), "倍速 speed", data.get("speed")])
    ws_summary.append(["部落数 tribes count", len(data.get("tribes", []))])

    headers = [
        "UnitId 兵种ID",
        "Name(EN) 英文名称",
        "名称(中文) Name(ZH)",
        "Type 类型",
        "Off 攻击",
        "Def_i 防御(步)",
        "Def_c 防御(骑)",
        "Speed 速度",
        "Cap 负重",
        "Cost 木 Lumber",
        "Cost 泥 Clay",
        "Cost 铁 Iron",
        "Cost 粮 Crop",
        "Upkeep 维护",
        "TrainTime 训练时间(s)",
        "TrainTimeFmt 训练时间",
        "ResearchTime 研究时间(s)",
        "ResearchTimeFmt 研究时间"
    ]

    for tribe in data.get("tribes", []):
        tid = int(tribe.get("tribeId"))
        en_label = tribe.get("tribeLabel") or f"tribe_{tid}"
        cn_label = CN_TRIBES.get(tid, en_label)
        title = f"{tid}-{cn_label}-{en_label}"
        ws = wb.create_sheet(title[:30])
        ws.append(headers)
        for u in tribe.get("units", []):
            uid = int(u.get("unitId"))
            en_name = u.get("name") or ""
            cn_name = CN_UNITS.get(tid, [])
            cn_name = cn_name[uid-1] if cn_name and uid-1 < len(cn_name) else en_name
            t = u.get("type")
            ws.append([
                uid,
                en_name,
                cn_name,
                CN_TYPE.get(str(t), str(t)),
                u.get("off"),
                u.get("def_i"),
                u.get("def_c"),
                u.get("speed"),
                u.get("cap"),
                (u.get("cost") or [None]*4)[0],
                (u.get("cost") or [None]*4)[1],
                (u.get("cost") or [None]*4)[2],
                (u.get("cost") or [None]*4)[3],
                u.get("cu"),
                u.get("time"),
                fmt_time(u.get("time")),
                u.get("rs_time"),
                fmt_time(u.get("rs_time"))
            ])
        for i in range(1, len(headers)+1):
            ws.column_dimensions[get_column_letter(i)].width = 16

    out_dir = os.path.dirname(out_path)
    if out_dir:
        os.makedirs(out_dir, exist_ok=True)
    wb.save(out_path)
    return out_path

def main():
    """命令行入口：导出默认路径的 xlsx"""
    base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "data", "troops_t4.6_1x.json"))
    out_path = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", "exports", "troops_t4.6_1x.xlsx"))
    path = export_excel(base, out_path)
    print(path)

if __name__ == "__main__":
    main()